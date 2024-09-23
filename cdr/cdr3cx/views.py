import os
import logging
from django.db.models.functions import Length
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
from django.db.models import Count
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.utils.dateparse import parse_datetime
from django.utils import timezone
from datetime import datetime, timedelta
from collections import Counter
from .project_numbers import COUNTRY_CODES
from .models import CallRecord, UserQuota
from django.shortcuts import render
from django.db.models.functions import Length
from django.db.models import Q,Sum, Count
from django.shortcuts import redirect
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET
logger = logging.getLogger(__name__)
from django.contrib.auth.decorators import login_required
from pytz import UTC
from .utils import get_country_from_number


@csrf_exempt
def receive_cdr(request):
    # Path to the file where records will be saved
    records_file_path = os.path.join(os.path.dirname(__file__), 'records.txt')

    # Log the request method, URL, and headers
    logger.info(f"Received request: {request.method} {request.build_absolute_uri()}")
    logger.info(f"Request headers: {request.headers}")

    if request.method == 'POST':
        # Log the full POST data
        raw_data = request.body.decode('utf-8')
        logger.info(f"Received raw data: {raw_data}")

        # Save raw data to records.txt for debugging
        with open(records_file_path, 'a') as f:
            f.write(f"{raw_data}\n")

        # Remove 'Call ' prefix if present
        if raw_data.startswith('Call '):
            raw_data = raw_data[5:]

        # Split the data
        cdr_data = raw_data.split(',')
        logger.info(f"Parsed CDR data: {cdr_data}")

        if len(cdr_data) < 3:
            logger.error("Insufficient data fields")
            return HttpResponse("Error: Insufficient data", status=400)

        # Extract and parse the data
        call_time_str, callee, caller = cdr_data[0], cdr_data[1], cdr_data[2]
        call_time_str = f"2024-07-26 {call_time_str}"  # Prepend the date for parsing

        # Parse the date and time
        try:
            call_time = parse_datetime(call_time_str)
            if call_time is None:
                raise ValueError(f"Failed to parse datetime from string: {call_time_str}")
        except Exception as e:
            logger.error(f"Error parsing call time: {e}")
            return HttpResponse(f"Error parsing call time: {e}", status=400)

        # Save to database
        try:
            call_record = CallRecord.objects.create(
                caller=caller.strip(),
                callee=callee.strip(),
                call_time=call_time,
                external_number=callee.strip()  # Assuming external_number is the same as callee
            )
            logger.info(f"Saved call record: {call_record}")
            return HttpResponse("CDR received and processed", status=200)
        except Exception as e:
            logger.error(f"Error saving call record: {e}")
            return HttpResponse("Error processing CDR", status=500)
    else:
        logger.error("Invalid request method")
        return HttpResponse("Error: Invalid request method", status=405)
    

def get_call_stats(queryset, time_period):
    now = timezone.now()
    
    if time_period == '1M':
        start_date = now - timedelta(days=30)
        date_trunc = TruncWeek
    elif time_period == '6M':
        start_date = now - timedelta(days=182)
        date_trunc = TruncMonth
    elif time_period == '1Y':
        start_date = now - timedelta(days=365)
        date_trunc = TruncMonth
    else:
        start_date = now - timedelta(days=30)
        date_trunc = TruncWeek

    call_stats = queryset.filter(call_time__range=[start_date, now]).annotate(
        date=TruncDate('call_time'),
        period=date_trunc('call_time')
    ).values('period').annotate(
        total_calls=Count('id'),
        local_calls=Count('id', filter=Q(callee__regex=r'^\d{4}$')),
        international_calls=Count('id', filter=Q(callee__startswith='00'))
    ).order_by('period')

    result = list(call_stats)
    logging.info(f"Call stats result: {result}")
    return result


from collections import Counter
from django.db.models.functions import Length

@login_required
def dashboard(request):
    now = timezone.now()
    time_period = request.GET.get('time_period', 'today')
    custom_date_range = request.GET.get('custom_date', '')

    if time_period == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = now
    elif time_period == '7d':
        start_date = now - timedelta(days=7)
        end_date = now
    elif time_period == '1m':
        start_date = now - timedelta(days=30)
        end_date = now
    elif time_period == '6m':
        start_date = now - timedelta(days=182)
        end_date = now
    elif time_period == '1y':
        start_date = now - timedelta(days=365)
        end_date = now
    elif time_period == 'custom' and custom_date_range:
        start_date_str, end_date_str = custom_date_range.split(" to ")
        start_date = timezone.make_aware(datetime.strptime(start_date_str, "%d %b, %Y"))
        end_date = timezone.make_aware(datetime.strptime(end_date_str, "%d %b, %Y").replace(hour=23, minute=59, second=59, microsecond=999999))
    else:
        start_date = now
        end_date = now

    # Filter call records by the selected time period
    call_records = CallRecord.objects.filter(call_time__range=[start_date, end_date])

    
    total_calls = call_records.count()
    total_external_calls = call_records.filter(
        Q(callee__regex=r'^\d{10}$') |
        Q(callee__regex=r'^\+966\d{9}$') |
        Q(callee__regex=r'^00966\d{9}$')
    ).count()
    total_international_calls = call_records.annotate(
        callee_length=Length('callee')
    ).filter(
        (
            Q(callee__startswith='+') |
            Q(callee__startswith='00')
        ) & ~(
            Q(callee__startswith='+966') |
            Q(callee__startswith='00966')
        ) & Q(callee_length__gt=11)
    ).count()
    total_national_mobile_calls = call_records.annotate(callee_length=Length('callee')).filter(callee__startswith='05', callee_length=10).count()
    total_national_calls = call_records.annotate(callee_length=Length('callee')).filter(callee__startswith='0', callee_length=9).count()
    total_local_calls = call_records.annotate(callee_length=Length('callee')).filter(
        callee_length__gt=4
    ).exclude(
        (
            Q(callee__startswith='+') |
            Q(callee__startswith='00')
        ) & ~(
            Q(callee__startswith='+966') |
            Q(callee__startswith='00966')
        ) & Q(callee_length__gt=11)
    ).count()

    # Calculate costs after filtering by time period
    total_call_cost = call_records.aggregate(Sum('total_cost'))['total_cost__sum'] or 0

    local_call_cost = call_records.annotate(callee_length=Length('callee')).filter(
        callee_length__gt=4
    ).exclude(
        (
            Q(callee__startswith='+') |
            Q(callee__startswith='00')
        ) & ~(
            Q(callee__startswith='+966') |
            Q(callee__startswith='00966')
        ) & Q(callee_length__gt=11)
    ).aggregate(Sum('total_cost'))['total_cost__sum'] or 0

    international_call_cost = call_records.annotate(
        callee_length=Length('callee')
    ).filter(
        (
            Q(callee__startswith='+') |
            Q(callee__startswith='00')
        ) & ~(
            Q(callee__startswith='+966') |
            Q(callee__startswith='00966')
        ) & Q(callee_length__gt=11)
    ).aggregate(Sum('total_cost'))['total_cost__sum'] or 0

    # Calculate top talking countries
    countries = [record.country for record in call_records if record.country not in ('Unknown', 'Internal Company Call')]
    top_talking_countries = Counter(countries).most_common(10)
    sort_by = request.GET.get('sort_by', 'amount')

    # New statistics for each caller
    caller_stats = call_records.filter(from_type='Extension').values('caller', 'from_dispname').annotate(
        total_calls=Count('id'),
        total_duration=Sum('duration'),
        external_calls=Count('id', filter=Q(to_type__in=['LineSet', 'Line'])),
        external_duration=Sum('duration', filter=Q(to_type__in=['LineSet', 'Line'])),
        company_calls=Count('id', filter=Q(to_type='Extension')),
        company_duration=Sum('duration', filter=Q(to_type='Extension')),
        total_cost=Sum('total_cost')
    )

    if sort_by == 'amount':
        caller_stats = caller_stats.order_by('-total_cost')
    elif sort_by == 'calls':
        caller_stats = caller_stats.order_by('-total_calls')
    else:
        caller_stats = caller_stats.order_by('-total_calls')  # Default sorting
    chart_time_period = request.GET.get('chart_time_period', '1M')
    call_stats = get_call_stats(CallRecord.objects, chart_time_period)

    context = {
        'call_records': call_records,
        'total_calls': total_calls,
        'total_external_calls': total_external_calls,
        'total_international_calls': total_international_calls,
        'total_national_mobile_calls': total_national_mobile_calls,
        'total_national_calls': total_national_calls,
        'total_local_calls': total_local_calls,
        'top_talking_countries': top_talking_countries,
        'time_period': time_period,
        'custom_date_range': custom_date_range,
        'caller_stats': caller_stats,
        'call_stats': call_stats,
        'chart_time_period': chart_time_period,
        'total_call_cost': total_call_cost,
        'local_call_cost': local_call_cost,
        'international_call_cost': international_call_cost,
        'sort_by': sort_by,

    }
    return render(request, 'cdr/dashboard.html', context)


def update_call_stats(request):
    chart_time_period = request.GET.get('chart_time_period', '1M')
    logging.info(f"Updating call stats for period: {chart_time_period}")
    
    call_stats = get_call_stats(CallRecord.objects, chart_time_period)
    logging.info(f"Call stats result: {call_stats}")
    
    return JsonResponse(call_stats, safe=False)

def get_call_stats(queryset, time_period):
    now = timezone.now()
    
    if time_period == '1M':
        start_date = now - timedelta(days=30)
    elif time_period == '6M':
        start_date = now - timedelta(days=182)
    elif time_period == '1Y':
        start_date = now - timedelta(days=365)
    else:
        raise ValueError(f"Invalid time period: {time_period}. Must be one of '1M', '6M', '1Y'")

    call_stats = queryset.filter(call_time__range=[start_date, now]).annotate(
        date=TruncDate('call_time'),
        period=TruncMonth('call_time')  # Using TruncMonth for all periods for simplicity
    ).values('period').annotate(
        total_calls=Count('id'),
        local_calls=Count('id', filter=Q(callee__regex=r'^\d{4}$')),
        international_calls=Count('id', filter=Q(callee__startswith='00'))
    ).order_by('period')

    result = list(call_stats)
    logging.info(f"Call stats result: {result}")
    return result


def update_country(request, record_id):
    if request.method == 'POST':
        country = request.POST.get('country')
        record = CallRecord.objects.get(id=record_id)
        record.country = country
        record.save()
    return redirect('dashboard')

def all_calls_view(request):
    search_query = request.GET.get('search', '')
    per_page = request.GET.get('per_page', 100)

    # Ensure per_page is an integer, defaulting to 100 if not a valid integer
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 100

    call_records = CallRecord.objects.all()

    if search_query:
        call_records = call_records.filter(caller__icontains=search_query) | call_records.filter(callee__icontains=search_query)

    paginator = Paginator(call_records, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'cdr/all_calls1.html', {'page_obj': page_obj, 'paginator': paginator, 'search_query': search_query, 'per_page': per_page})



def outgoingExtCalls(request):
    search_query = request.GET.get('search', '')
    per_page = request.GET.get('per_page', 100)

    # Ensure per_page is an integer, defaulting to 100 if not a valid integer
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 100

    call_records = CallRecord.objects.filter(Q(to_type="LineSet") | Q(to_type="Line"))

    if search_query:
        call_records = call_records.filter(caller__icontains=search_query) | call_records.filter(callee__icontains=search_query)

    paginator = Paginator(call_records, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'cdr/outgoingExtCalls.html', {'page_obj': page_obj, 'paginator': paginator, 'search_query': search_query, 'per_page': per_page})

def incomingCalls(request):
    search_query = request.GET.get('search', '')
    per_page = request.GET.get('per_page', 100)

    # Ensure per_page is an integer, defaulting to 100 if not a valid integer
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 100

    call_records = CallRecord.objects.filter(from_type="Line")

    if search_query:
        call_records = call_records.filter(Q(caller__icontains=search_query) | Q(callee__icontains=search_query))

    paginator = Paginator(call_records, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'cdr/incomingCalls.html', {'page_obj': page_obj, 'paginator': paginator, 'search_query': search_query, 'per_page': per_page})

def outgoingInternationalCalls(request):
    search_query = request.GET.get('search', '')
    per_page = request.GET.get('per_page', 100)

    # Ensure per_page is an integer, defaulting to 100 if not a valid integer
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 100

    # Define the filtering logic
    call_records = CallRecord.objects.filter(
        Q(callee__regex=r'^\+[^9]') | 
        Q(callee__regex=r'^\+9[0-8]') | 
        Q(callee__regex=r'^00[^9]') | 
        Q(callee__regex=r'^009[0-8]')
    ).exclude(
        Q(callee__startswith='+966') | Q(callee__startswith='00966')
    )

    if search_query:
        call_records = call_records.filter(Q(caller__icontains=search_query) | Q(callee__icontains=search_query))

    paginator = Paginator(call_records, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'cdr/outgoingInternationalCalls.html', {'page_obj': page_obj, 'paginator': paginator, 'search_query': search_query, 'per_page': per_page})

def local_calls_view(request):
    call_records = CallRecord.objects.annotate(callee_length=Length('callee')).filter(callee_length=4)
    return render(request, 'cdr/local_calls.html', {'call_records': call_records})

def national_calls_view(request):
    call_records = CallRecord.objects.filter(callee__startswith='0').exclude(callee__startswith='00')
    return render(request, 'cdr/national_calls.html', {'call_records': call_records})

def international_calls_view(request):
    call_records = CallRecord.objects.annotate(callee_length=Length('callee')).filter(callee__startswith='00', callee_length__gt=10)
    return render(request, 'cdr/international_calls.html', {'call_records': call_records})

def home(request):
    return render(request, 'home/home.html')
def aboutus(request):
    return render(request, 'home/about_us.html')

from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
from .models import CallRecord

from django.db.models.functions import Substr


from django.db.models.functions import Substr, Length
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET

from django.utils import timezone
from datetime import timedelta



@require_GET
def get_caller_record(request):
    caller_number = request.GET.get('caller_number')
    if caller_number:
        try:
            # Using slicing to match the last 9 digits of from_no with caller_number
            caller_last_9 = caller_number[-9:]
            print(f"Caller last 9 digits: {caller_last_9}")
            
            # Get the current time and calculate 48 hours ago
            now = timezone.now()
            forty_eight_hours_ago = now - timedelta(hours=36)
            
            # Annotate and filter the records
            record = CallRecord.objects.annotate(
                last_9_from_no=Substr('callee', Length('callee') - 8, 9)
            ).filter(
                last_9_from_no=caller_last_9,
                call_time__gte=forty_eight_hours_ago  # Filter for calls within the last 48 hours
            ).order_by('-call_time').first()
            
            if record:
                # Check if the caller is a 4-digit extension
                if record.caller.isdigit() and len(record.caller) == 4:
                    return HttpResponse(record.caller, status=200)
                else:
                    return HttpResponse("", status=200)
            else:
                print("No record found within the last 48 hours")
                return HttpResponse("", status=404)
        except Exception as e:
            print(f"Error: {e}")
            return JsonResponse({'error': str(e)}, status=500)
    else:
        print("caller_number parameter is required")
        return JsonResponse({'error': 'caller_number parameter is required'}, status=400)
    
from django.utils.text import slugify

def country_specific_calls_view(request, country_slug):
    search_query = request.GET.get('search', '')
    per_page = request.GET.get('per_page', 100)

    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 100

    # Convert slug back to country name
    country = country_slug.replace('-', ' ').title()

    call_records = CallRecord.objects.filter(country=country)

    if search_query:
        call_records = call_records.filter(Q(caller__icontains=search_query) | Q(callee__icontains=search_query))

    paginator = Paginator(call_records, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'paginator': paginator,
        'search_query': search_query,
        'per_page': per_page,
        'country': country,
    }

    return render(request, 'cdr/country_specific_calls.html', context)

from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q
from .models import CallRecord
from django.db.models import Sum, Count
from django.utils import timezone
from dateutil.relativedelta import relativedelta

def caller_calls_view(request, caller_number):
    search_query = request.GET.get('search', '')
    per_page = request.GET.get('per_page', 100)
    date_filter = request.GET.get('date_filter', '1M')
    custom_date_range = request.GET.get('custom_date', '')
    export_excel = request.GET.get('export_excel', 'false') == 'true'

    print(f"Debug: date_filter = {date_filter}")
    print(f"Debug: custom_date_range = {custom_date_range}")

    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 100

    call_records = CallRecord.objects.filter(caller=caller_number)
    

    # Apply date filter
    now = timezone.now()
    if custom_date_range:
        try:
            start_date_str, end_date_str = custom_date_range.split(" to ")
            start_date = timezone.make_aware(datetime.strptime(start_date_str.strip(), "%d %b, %Y"))
            end_date = timezone.make_aware(datetime.strptime(end_date_str.strip(), "%d %b, %Y").replace(hour=23, minute=59, second=59))
            call_records = call_records.filter(call_time__range=[start_date, end_date])
            date_filter = 'custom'  # Set date_filter to 'custom' when using custom range
            print(f"Debug: Custom date range applied. Start: {start_date}, End: {end_date}")
        except ValueError as e:
            print(f"Debug: Error parsing custom date range: {e}")
            
    elif date_filter == 'ALL':
        pass
    elif date_filter == 'today':
        call_records = call_records.filter(call_time__date=now.date())
    elif date_filter == '7d':
        call_records = call_records.filter(call_time__gte=now - timedelta(days=7))
    elif date_filter == '1M':
        call_records = call_records.filter(call_time__gte=now - relativedelta(months=1))
    elif date_filter == '6M':
        call_records = call_records.filter(call_time__gte=now - relativedelta(months=6))
    elif date_filter == '1Y':
        call_records = call_records.filter(call_time__gte=now - relativedelta(years=1))

    print(f"Debug: Filtered call_records count: {call_records.count()}")
   

    if search_query:
        call_records = call_records.filter(
            Q(callee__icontains=search_query) | 
            Q(call_time__icontains=search_query)
        )
       

    # Calculate statistics
    total_calls = call_records.count()
    total_cost = call_records.aggregate(Sum('total_cost'))['total_cost__sum'] or 0
    total_duration = call_records.aggregate(Sum('duration'))['duration__sum'] or 0

    total_local_calls = call_records.annotate(callee_length=Length('callee')).filter(
        callee_length__gt=4
    ).exclude(
        (
            Q(callee__startswith='+') |
            Q(callee__startswith='00')
        ) & ~(
            Q(callee__startswith='+966') |
            Q(callee__startswith='00966')
        ) & Q(callee_length__gt=11)
    ).count()

    local_call_cost = call_records.annotate(callee_length=Length('callee')).filter(
        callee_length__gt=4
    ).exclude(
        (
            Q(callee__startswith='+') |
            Q(callee__startswith='00')
        ) & ~(
            Q(callee__startswith='+966') |
            Q(callee__startswith='00966')
        ) & Q(callee_length__gt=11)
    ).aggregate(Sum('total_cost'))['total_cost__sum'] or 0

    total_incoming_calls = call_records.filter(from_type="Line").count()
    incoming_call_cost = call_records.filter(from_type="Line").aggregate(Sum('total_cost'))['total_cost__sum'] or 0

    total_international_calls = call_records.annotate(
        callee_length=Length('callee')
    ).filter(
        (
            Q(callee__startswith='+') |
            Q(callee__startswith='00')
        ) & ~(
            Q(callee__startswith='+966') |
            Q(callee__startswith='00966')
        ) & Q(callee_length__gt=11)
    ).count()

    international_call_cost = call_records.annotate(
        callee_length=Length('callee')
    ).filter(
        (
            Q(callee__startswith='+') |
            Q(callee__startswith='00')
        ) & ~(
            Q(callee__startswith='+966') |
            Q(callee__startswith='00966')
        ) & Q(callee_length__gt=11)
    ).aggregate(Sum('total_cost'))['total_cost__sum'] or 0

    if export_excel:
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = f"Calls for {caller_number}"

        # Add headers
        headers = ['Call Time', 'Callee', 'Duration', 'Call Type', 'Cost']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Add data
        for row, call in enumerate(call_records, start=2):
            ws.cell(row=row, column=1, value=call.call_time.strftime("%d %b, %Y %H:%M:%S"))
            ws.cell(row=row, column=2, value=call.callee)
            # Handle potential None values for duration
            duration_str = str(timedelta(seconds=call.duration)) if call.duration is not None else "N/A"
            ws.cell(row=row, column=3, value=duration_str)
            ws.cell(row=row, column=4, value=call.to_type)
            ws.cell(row=row, column=5, value=f"{call.total_cost:.2f} SAR")

        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=calls_for_{caller_number}.xlsx'

        wb.save(response)
        return response
    
    # Paginate the filtered results
    paginator = Paginator(call_records, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'paginator': paginator,
        'search_query': search_query,
        'per_page': per_page,
        'caller_number': caller_number,
        'date_filter': date_filter,
        'custom_date_range': custom_date_range,
        'total_cost': total_cost,
        'total_calls': total_calls,
        'total_duration': total_duration,
        'total_local_calls': total_local_calls,
        'local_call_cost': local_call_cost,
        'total_incoming_calls': total_incoming_calls,
        'incoming_call_cost': incoming_call_cost,
        'total_international_calls': total_international_calls,
        'international_call_cost': international_call_cost,
    }

    
   

    return render(request, 'cdr/caller_calls.html', context)


from django.db.models import Count, Sum
from django.db.models.functions import Substr

def call_record_summary_view(request):
    # Grouping records by category and pattern prefix
    summary = CallRecord.objects.annotate(
        pattern_prefix=Substr('callee', 1, 3)  # Adjust the substring length to match your pattern
    ).values(
        'call_category',
        'pattern_prefix'
    ).annotate(
        count=Count('id'),
        total_cost=Sum('total_cost')
    ).order_by('call_category', 'pattern_prefix')

    context = {
        'summary': summary,
    }
    return render(request, "cdr/call_record_summary.html", context)


from django.core.mail import send_mail
from django.conf import settings
from decimal import Decimal

def check_balance_and_send_email():
    # Get all UserQuota objects
    user_quotas = UserQuota.objects.all()

    for user_quota in user_quotas:
        if user_quota.quota:
            # Calculate the percentage of remaining balance
            total_quota = user_quota.quota.amount
            remaining_balance = user_quota.remaining_balance
            remaining_percentage = (remaining_balance / total_quota) * 100

            # Check if remaining balance is 50% or less
            if remaining_percentage <= 50:
                # Prepare email content
                subject = f"Low Balance Alert for Extension {user_quota.extension}"
                message = f"""
                Dear Administrator,

                The remaining balance for extension {user_quota.extension} is now at {remaining_percentage:.2f}% of its total quota.

                Total Quota: {total_quota}
                Remaining Balance: {remaining_balance}

                Please take necessary actions.

                Best regards,
                Your System
                """
                from_email = settings.DEFAULT_FROM_EMAIL
                recipient_list = ['khuram2025@gmail.com']  # Add any other recipients here

                # Send email
                send_mail(
                    subject,
                    message,
                    from_email,
                    recipient_list,
                    fail_silently=False,
                )

                print(f"Email sent for extension {user_quota.extension}")


from django.shortcuts import render
from django.db.models import Count, Sum, F, Q
from django.http import HttpResponse, HttpResponseBadRequest

from django.utils import timezone
from datetime import datetime, timedelta
from .models import CallRecord
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def get_date_range(request):
    now = timezone.now()
    time_period = request.GET.get('time_period', 'today')
    custom_date_range = request.GET.get('custom_date', '')

    if time_period == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = now
    elif time_period == '7d':
        start_date = now - timedelta(days=7)
        end_date = now
    elif time_period == '1m':
        start_date = now - timedelta(days=30)
        end_date = now
    elif time_period == '6m':
        start_date = now - timedelta(days=182)
        end_date = now
    elif time_period == '1y':
        start_date = now - timedelta(days=365)
        end_date = now
    elif time_period == 'custom' and custom_date_range:
        start_date_str, end_date_str = custom_date_range.split(" to ")
        start_date = timezone.make_aware(datetime.strptime(start_date_str, "%d %b, %Y"))
        end_date = timezone.make_aware(datetime.strptime(end_date_str, "%d %b, %Y").replace(hour=23, minute=59, second=59, microsecond=999999))
    else:
        start_date = now
        end_date = now

    return start_date, end_date, time_period, custom_date_range



def calculate_call_stats(call_records):
    total_calls = call_records.count()
    total_call_cost = call_records.aggregate(Sum('total_cost'))['total_cost__sum'] or 0

    total_local_calls = call_records.filter(
        Q(callee__regex=r'^\d{10}$') |
        Q(callee__regex=r'^\+966\d{9}$') |
        Q(callee__regex=r'^00966\d{9}$')
    ).count()
    local_call_cost = call_records.filter(
        Q(callee__regex=r'^\d{10}$') |
        Q(callee__regex=r'^\+966\d{9}$') |
        Q(callee__regex=r'^00966\d{9}$')
    ).aggregate(Sum('total_cost'))['total_cost__sum'] or 0

    total_incoming_calls = call_records.filter(to_type='Extension').count()
    
    total_international_calls = call_records.exclude(
        Q(callee__regex=r'^\d{10}$') |
        Q(callee__regex=r'^\+966\d{9}$') |
        Q(callee__regex=r'^00966\d{9}$')
    ).count()
    international_call_cost = call_records.exclude(
        Q(callee__regex=r'^\d{10}$') |
        Q(callee__regex=r'^\+966\d{9}$') |
        Q(callee__regex=r'^00966\d{9}$')
    ).aggregate(Sum('total_cost'))['total_cost__sum'] or 0

    return {
        'total_calls': total_calls,
        'total_call_cost': total_call_cost,
        'total_local_calls': total_local_calls,
        'local_call_cost': local_call_cost,
        'total_incoming_calls': total_incoming_calls,
        'total_international_calls': total_international_calls,
        'international_call_cost': international_call_cost,
    }

def get_top_extensions_data(request, sort_by):
    start_date, end_date, time_period, custom_date_range = get_date_range(request)
    limit = request.GET.get('limit', 20)
    try:
        limit = int(limit)
    except ValueError:
        limit = 20

    # Define valid sorting options
    valid_sort_options = {
        'total_calls': '-total_calls',
        'talk_time': '-total_duration',
        'cost': '-total_cost'
    }

    # Use the provided sort_by if valid, otherwise default to 'total_calls'
    sort_field = valid_sort_options.get(sort_by, '-total_calls')

    top_extensions = CallRecord.objects.filter(
        call_time__range=[start_date, end_date],
        to_type='Line'
    ).values('caller').annotate(
        total_calls=Count('id'),
        total_duration=Sum('duration'),
        total_cost=Sum('total_cost')
    ).order_by(sort_field)[:limit]

    top_callers = [ext['caller'] for ext in top_extensions]
    call_records = CallRecord.objects.filter(
        call_time__range=[start_date, end_date],
        caller__in=top_callers,
        to_type='Line'
    )

    call_stats = calculate_call_stats(call_records)

    return {
        'top_extensions': top_extensions,
        'title': f'Top {limit} Extensions by {sort_by.replace("_", " ").title()} (Line Calls)',
        'time_period': time_period,
        'custom_date_range': custom_date_range,
        'limit': limit,
        'start_date': start_date,
        'end_date': end_date,
        **call_stats,
    }



def top_extensions(request):
    context = get_top_extensions_data(request, 'total_calls')
    return render(request, 'cdr/top_extensions.html', context)

def top_extensions_talk_time(request):
    context = get_top_extensions_data(request, 'total_duration')
    return render(request, 'cdr/top_extensions.html', context)

def top_extensions_cost(request):
    context = get_top_extensions_data(request, 'total_cost')
    return render(request, 'cdr/top_extensions.html', context)


def generate_excel_report(request):
    # Map URL names to sort_by values
    url_to_sort_map = {
        'top_extensions': 'total_calls',
        'top_extensions_talk_time': 'talk_time',
        'top_extensions_cost': 'cost'
    }

    # Get the current URL name
    current_url_name = request.resolver_match.url_name

    # Determine the sort_by value
    sort_by = url_to_sort_map.get(current_url_name, 'total_calls')

    try:
        data = get_top_extensions_data(request, sort_by)
    except Exception as e:
        return HttpResponseBadRequest(f"Error generating report: {str(e)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Top Extensions Report"

    # Add title
    ws['A1'] = data['title']
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')

    # Add date range
    ws['A2'] = f"Date Range: {data['start_date'].strftime('%Y-%m-%d')} to {data['end_date'].strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(size=12)
    ws.merge_cells('A2:D2')

    # Add headers
    headers = ['Extension', 'Total Calls', 'Total Duration', 'Total Cost']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Add data
    for row, ext in enumerate(data['top_extensions'], start=5):
        ws.cell(row=row, column=1, value=ext['caller'])
        ws.cell(row=row, column=2, value=ext['total_calls'])
        ws.cell(row=row, column=3, value=str(timedelta(seconds=ext['total_duration'])))
        ws.cell(row=row, column=4, value=f"{ext['total_cost']:.2f} SAR")

    # Adjust column widths
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{sort_by}_top_extensions_report.xlsx"'

    wb.save(response)

    return response